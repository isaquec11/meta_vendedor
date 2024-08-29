import pandas as pd
from openpyxl import load_workbook

from twilio.rest import Client

# Find your Account SID and Auth Token at twilio.com/console
# and set the environment variables. See http://twil.io/secure
account_sid = 'ACee0ab7c1d0173b11724298f14aef1f66'
auth_token = 'a76784988d062bc85ef6edcf6547c550'
client = Client(account_sid, auth_token)


# Vendedor que bateu 50.000 em vendas

lista_meses = ['janeiro', 'fevereiro', 'março', 'abril', 'maio', 'junho']

# Abrir os arquivos em excel
for mes in lista_meses:
    # Abrir a planilha do mês
    planilha = load_workbook(f'{mes}.xlsx')
    pagina = planilha.active  
    for linha in pagina.iter_rows(min_row=2, values_only=True):
        vendedor = linha[0]
        vendas = linha[1]
        if vendas > 55000:
            print(f'No mês {mes}, alguém cumpriu a meta. Vendedor: {vendedor}, vendas em R$: {vendas}')
            
            # Enviar uma mensagem via Twilio
            message = client.messages.create(
                body=f'No mês {mes}, alguém cumpriu a meta. Vendedor: {vendedor}, vendas em R$: {vendas}',
                from_='+12563304386',
                to='+5585997379431'
            )
            print(message.sid)




